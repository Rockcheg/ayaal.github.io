#!/usr/bin/env python3
import argparse
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook

SHEET_NAME = 'Сайт_экспорт'

ALL_FIELDS = [
    'id', 'name', 'category', 'subcategory', 'price', 'deliveryCost',
    'weight', 'sku', 'availability', 'image', 'images', 'description',
    'testedByUs', 'supplierType'
]

REQUIRED_FIELDS = {
    'id', 'name', 'category', 'subcategory', 'price', 'deliveryCost',
    'weight', 'sku', 'availability', 'image', 'description'
}

TRUE_VALUES = {'true', 'да', '1', 'yes', 'y'}
VALID_SUPPLIER_TYPES = {'factory', 'supplier'}


def to_text(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def to_int(value, field_name: str, row_num: int, warnings: List[str]):
    raw = to_text(value)
    if not raw:
        return None
    try:
        return int(float(raw.replace(' ', '').replace(',', '.')))
    except ValueError:
        warnings.append(f"WARNING: row {row_num}: field '{field_name}' has non-numeric value: {raw!r}")
        return None


def parse_bool(value) -> bool:
    raw = to_text(value).lower()
    return raw in TRUE_VALUES


def split_images(value) -> List[str]:
    raw = to_text(value)
    if not raw:
        return []
    return [item.strip() for item in raw.split(';') if item.strip()]


def escape_js_string(value: str) -> str:
    return value.replace('\\', '\\\\').replace("'", "\\'")


def build_case_sensitive_index(root: Path) -> Dict[Tuple[str, ...], str]:
    index = {}
    for p in root.rglob('*'):
        if p.is_file():
            rel_parts = p.relative_to(root).parts
            key = tuple(part.lower() for part in rel_parts)
            index[key] = '/'.join(rel_parts)
    return index


def check_image_path(path_str: str, repo_root: Path, case_index: Dict[Tuple[str, ...], str], row_num: int, field_name: str, warnings: List[str]):
    normalized = path_str.replace('\\', '/').strip()
    if not normalized:
        return
    key = tuple(part for part in normalized.split('/') if part)
    lower_key = tuple(part.lower() for part in key)

    if lower_key not in case_index:
        warnings.append(f"WARNING: row {row_num}: field '{field_name}' image not found: {normalized}")
        return

    actual_case_path = case_index[lower_key]
    expected_path = '/'.join(key)
    if actual_case_path != expected_path:
        warnings.append(
            f"WARNING: row {row_num}: field '{field_name}' case mismatch: '{normalized}' vs actual '{actual_case_path}'"
        )


def format_product(product: dict, indent: str = '  ') -> List[str]:
    lines = [f"{indent}{{"]
    field_order = [
        'id', 'name', 'category', 'subcategory', 'price', 'deliveryCost',
        'weight', 'sku', 'availability', 'testedByUs', 'supplierType', 'image', 'images', 'description'
    ]

    rendered = []
    for field in field_order:
        if field not in product:
            continue
        value = product[field]
        if isinstance(value, bool):
            rendered.append(f"{indent}  {field}: {'true' if value else 'false'},")
        elif isinstance(value, int):
            rendered.append(f"{indent}  {field}: {value},")
        elif isinstance(value, list):
            if not value:
                continue
            rendered.append(f"{indent}  {field}: [")
            for item in value:
                rendered.append(f"{indent}    '{escape_js_string(item)}',")
            rendered.append(f"{indent}  ],")
        else:
            rendered.append(f"{indent}  {field}: '{escape_js_string(str(value))}',")

    if rendered:
        rendered[-1] = rendered[-1].rstrip(',')
    lines.extend(rendered)
    lines.append(f"{indent}}}")
    return lines


def main():
    parser = argparse.ArgumentParser(description='Export products from Excel into js/products.js')
    parser.add_argument('--excel', required=True, help='Path to source .xlsx file')
    parser.add_argument('--out', required=True, help='Path to output js file')
    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parents[1]
    excel_path = (repo_root / args.excel).resolve()
    out_path = (repo_root / args.out).resolve()

    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit('Excel sheet is empty')

    headers = [to_text(h) for h in rows[0]]
    header_map = {h: idx for idx, h in enumerate(headers) if h}

    missing_columns = [f for f in ALL_FIELDS if f not in header_map]
    warnings: List[str] = []
    if missing_columns:
        warnings.append('WARNING: missing columns in header: ' + ', '.join(missing_columns))

    case_index = build_case_sensitive_index(repo_root)
    products = []

    for row_offset, row in enumerate(rows[1:], start=2):
        raw = {field: row[header_map[field]] if field in header_map and header_map[field] < len(row) else None for field in ALL_FIELDS}
        if all(to_text(v) == '' for v in raw.values()):
            continue

        for field in REQUIRED_FIELDS:
            if to_text(raw.get(field)) == '':
                warnings.append(f"WARNING: row {row_offset}: required field '{field}' is empty")

        product = {}
        product_id = to_int(raw.get('id'), 'id', row_offset, warnings)
        if product_id is not None:
            product['id'] = product_id

        for text_field in ['name', 'category', 'subcategory', 'weight', 'sku', 'availability', 'image', 'description']:
            text_val = to_text(raw.get(text_field))
            if text_val:
                product[text_field] = text_val

        for int_field in ['price', 'deliveryCost']:
            int_val = to_int(raw.get(int_field), int_field, row_offset, warnings)
            if int_val is not None:
                product[int_field] = int_val

        if parse_bool(raw.get('testedByUs')):
            product['testedByUs'] = True

        supplier_type = to_text(raw.get('supplierType')).lower()
        if supplier_type in VALID_SUPPLIER_TYPES:
            product['supplierType'] = supplier_type

        image_list = split_images(raw.get('images'))
        if image_list:
            product['images'] = image_list

        image_main = to_text(raw.get('image'))
        if image_main:
            check_image_path(image_main, repo_root, case_index, row_offset, 'image', warnings)
        for img in image_list:
            check_image_path(img, repo_root, case_index, row_offset, 'images', warnings)

        products.append(product)

    output_lines = ['window.products = [']
    for idx, product in enumerate(products):
        product_lines = format_product(product)
        if idx < len(products) - 1:
            product_lines[-1] += ','
        output_lines.extend(product_lines)
    output_lines.append('];')
    output_text = '\n'.join(output_lines) + '\n'

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(output_text, encoding='utf-8')

    for warning in warnings:
        print(warning)
    print(f"Generated {out_path} from {excel_path} ({len(products)} products)")


if __name__ == '__main__':
    main()
